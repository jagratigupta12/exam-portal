"""
Parse questions from Excel (.xlsx), Word (.docx), or plain text (.txt)

EXCEL FORMAT — 9 columns, Row 1 = headers, data from Row 2:
  A: MCQ Question
  B: Option A
  C: Option B
  D: Option C
  E: Option D
  F: Correct Option (A/B/C/D)
  G: Marks (default 1)
  H: One-Word Question (for Test 2)
  I: One-Word Answer
"""

def parse_excel(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl not installed. Please run this command in PyCharm terminal:\n"
            "pip install openpyxl --break-system-packages\n"
            "OR: pip install openpyxl"
        )

    wb = load_workbook(file_obj)
    ws = wb.active
    questions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[0] or str(row[0]).strip() == '':
            continue
        try:
            q = {
                'question_text': str(row[0]).strip(),
                'option_a':      str(row[1]).strip() if row[1] is not None else '',
                'option_b':      str(row[2]).strip() if row[2] is not None else '',
                'option_c':      str(row[3]).strip() if row[3] is not None else '',
                'option_d':      str(row[4]).strip() if row[4] is not None else '',
                'correct_option':str(row[5]).strip().upper() if row[5] is not None else 'A',
                'marks':         int(float(str(row[6]))) if row[6] is not None else 1,
                'oneword_question': str(row[7]).strip() if row[7] is not None else '',
                'oneword_answer':   str(row[8]).strip().lower() if row[8] is not None else '',
            }
            # Validate correct option
            if q['correct_option'] not in ['A', 'B', 'C', 'D']:
                q['correct_option'] = 'A'
            if q['question_text'] and q['option_a']:
                questions.append(q)
        except Exception as e:
            continue  # Skip bad rows silently

    return questions


def parse_word(file_obj):
    try:
        from docx import Document
    except ImportError:
        raise ImportError(
            "python-docx not installed. Run: pip install python-docx"
        )
    doc = Document(file_obj)
    full_text = '\n'.join(p.text for p in doc.paragraphs)
    return parse_text(full_text)


def parse_text(text):
    import re
    questions = []
    blocks = re.split(r'\n\s*---+\s*\n', text.strip())
    for block in blocks:
        lines = [l.strip() for l in block.strip().splitlines() if l.strip()]
        q = {}
        for line in lines:
            for prefix, key in [
                ('Q:', 'question_text'), ('A:', 'option_a'), ('B:', 'option_b'),
                ('C:', 'option_c'),      ('D:', 'option_d'), ('ANS:', 'correct_option'),
                ('MARKS:', 'marks'),     ('OW:', 'oneword_question'), ('OWA:', 'oneword_answer'),
            ]:
                if line.upper().startswith(prefix):
                    q[key] = line[len(prefix):].strip()
                    break
        if 'question_text' in q and 'option_a' in q:
            q.setdefault('option_b', '')
            q.setdefault('option_c', '')
            q.setdefault('option_d', '')
            q.setdefault('correct_option', 'A')
            q.setdefault('marks', 1)
            q.setdefault('oneword_question', '')
            q.setdefault('oneword_answer', '')
            try:
                q['marks'] = int(q['marks'])
            except Exception:
                q['marks'] = 1
            q['correct_option'] = str(q['correct_option']).upper()
            questions.append(q)
    return questions
