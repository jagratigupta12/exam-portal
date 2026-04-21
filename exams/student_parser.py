"""
Parse student list from Excel, Word, or Text file.

EXCEL FORMAT (.xlsx):
  Col A: Student Name
  Col B: Enrollment No
  Col C: Email (optional)

WORD/TEXT FORMAT (.docx / .txt):
  Rahul Sharma, 2021CS001
  Priya Singh, 2021CS002, priya@email.com
"""

def parse_excel(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Run: pip install openpyxl")
    
    wb = load_workbook(file_obj)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name       = str(row[0]).strip() if row[0] else ''
        enrollment = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        email      = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if name:
            students.append({'name': name, 'enrollment': enrollment, 'email': email})
    return students


def parse_word(file_obj):
    try:
        from docx import Document
    except ImportError:
        raise ImportError("Run: pip install python-docx")
    doc = Document(file_obj)
    text = '\n'.join(p.text for p in doc.paragraphs)
    return parse_text(text)


def parse_text(text):
    students = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(',')]
        name       = parts[0] if len(parts) > 0 else ''
        enrollment = parts[1] if len(parts) > 1 else ''
        email      = parts[2] if len(parts) > 2 else ''
        if name:
            students.append({'name': name, 'enrollment': enrollment, 'email': email})
    return students
